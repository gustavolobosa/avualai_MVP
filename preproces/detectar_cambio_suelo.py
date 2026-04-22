# detectar_cambio_suelo.py
# Detecta predios agrícolas que cambiaron tipo de suelo de secano a riego entre 2019 y 2024.
# No modifica la base de datos. Genera reportes CSV.

import pandas as pd
import numpy as np
from pathlib import Path

# ========================
# CONFIG
# ========================
DIR_CONSTRUCCIONES = Path(__file__).resolve().parent.parent / 'datos' / 'construcciones'
DIR_PREDIOS = Path(__file__).resolve().parent.parent / 'datos' / 'predios'
OUTPUT_DIR = Path(__file__).resolve().parent

ARCHIVO_AL_2019 = DIR_CONSTRUCCIONES / 'BRORGA2441AL_NAC_2019_1'
ARCHIVO_AL_2024 = DIR_CONSTRUCCIONES / 'BRORGA2441AL_NAC_2024_1'
ARCHIVO_PREDIOS_2019 = DIR_PREDIOS / 'BRTMPNACROL_NAC_2019_1'
ARCHIVO_PREDIOS_2024 = DIR_PREDIOS / 'BRTMPNACROL_NAC_2024_1'

# ========================
# CONSTANTES
# ========================
SECANO_CODES = {
    '1', '2', '3', '4', '5', '6', '7', '8',
    '20', '21', '22', '23', '24', '25', '26', '27', '28'
}
RIEGO_CODES = {'1R', '2R', '3R', '4R'}

COLUMNAS_12 = [
    'cod_comuna', 'manzana_actual', 'predio_actual', 'cod_suelo', 'sup_suelo',
    'correlativo_linea_constru', 'cod_material', 'cod_calidad', 'sup_constru',
    'cod_destino', 'cod_condicion_especial', 'num_pisos'
]
COLUMNAS_11 = COLUMNAS_12[:-1]  # sin num_pisos (archivo 2019)

# Predios: fixed-width format
PREDIOS_COLSPECS = [
    (0, 5),     # comuna_actual
    (5, 9),     # anio
    (9, 10),    # semestre
    (10, 11),   # indicador_de_aseo
    (17, 57),   # direccion_predial
    (57, 62),   # manzana_actual
    (62, 67),   # predio_actual
    (67, 68),   # cod_serie
    (68, 81),   # cuota_trimestral
    (81, 96),   # avaluo_tot
    (96, 111),  # avaluo_ex
    (111, 115), # anio_termino_ex
    (115, 116), # cod_ubi
    (116, 117)  # cod_destino
]
PREDIOS_COLNAMES = [
    'comuna_actual', 'anio', 'semestre', 'indicador_de_aseo', 'direccion_predial',
    'manzana_actual', 'predio_actual', 'cod_serie', 'cuota_trimestral',
    'avaluo_tot', 'avaluo_ex', 'anio_termino_ex', 'cod_ubi', 'cod_destino'
]


def clasificar_suelo(cod):
    cod = str(cod).strip()
    if cod in RIEGO_CODES:
        return 'riego'
    if cod in SECANO_CODES:
        return 'secano'
    return 'otro'


def detectar_num_columnas(filepath):
    """Lee la primera línea para determinar si tiene 11 o 12 columnas."""
    with open(filepath, 'r', encoding='latin1') as f:
        primera = f.readline()
    n_pipes = primera.count('|')
    return n_pipes + 1  # n separadores = n-1 campos? No: n pipes = n+1 si no hay trailing
    # En realidad estos archivos tienen trailing pipe, así que n_pipes = n_columnas


def cargar_suelos(filepath):
    """Carga un archivo AL y retorna solo las líneas de suelo (correlativo == 0)."""
    print(f"  Cargando {filepath.name}...")

    # Detectar columnas
    with open(filepath, 'r', encoding='latin1') as f:
        primera = f.readline()
    n_campos = len(primera.rstrip('\n').rstrip('\r').split('|'))
    # El archivo puede tener trailing pipe que genera un campo vacío extra
    if n_campos >= 12:
        columnas = COLUMNAS_12
    else:
        columnas = COLUMNAS_11

    df = pd.read_csv(
        filepath, sep='|', names=columnas, dtype=str,
        index_col=False, encoding='latin1'
    )
    # Si se leyeron más columnas de las esperadas (trailing pipe), eliminar extras
    for col in df.columns:
        if col not in columnas:
            df = df.drop(columns=[col])

    # Limpiar
    df['cod_suelo'] = df['cod_suelo'].astype(str).str.strip()
    df['correlativo_linea_constru'] = pd.to_numeric(
        df['correlativo_linea_constru'], errors='coerce'
    ).fillna(-1).astype(int)
    df['sup_suelo'] = pd.to_numeric(df['sup_suelo'], errors='coerce').fillna(0)

    # Construir clave_predio
    df['cod_comuna'] = df['cod_comuna'].astype(str).str.strip().str.zfill(5)
    df['manzana_actual'] = pd.to_numeric(df['manzana_actual'], errors='coerce').astype('Int64')
    df['predio_actual'] = pd.to_numeric(df['predio_actual'], errors='coerce').astype('Int64')
    df['clave_predio'] = (
        df['cod_comuna'] + '-' +
        df['manzana_actual'].astype(str) + '-' +
        df['predio_actual'].astype(str)
    )

    # Filtrar solo líneas de suelo (correlativo == 0, cod_suelo no vacío)
    df_suelo = df[
        (df['correlativo_linea_constru'] == 0) &
        (df['cod_suelo'] != '') &
        (df['cod_suelo'] != 'nan')
    ].copy()

    # Clasificar
    df_suelo['tipo_suelo'] = df_suelo['cod_suelo'].apply(clasificar_suelo)

    # Dedup por (clave_predio, cod_suelo) keep last
    df_suelo = df_suelo.drop_duplicates(
        subset=['clave_predio', 'cod_suelo'], keep='last'
    )

    print(f"    Total líneas: {len(df):,}")
    print(f"    Líneas de suelo: {len(df_suelo):,}")
    print(f"    Predios únicos: {df_suelo['clave_predio'].nunique():,}")

    return df_suelo[['clave_predio', 'cod_comuna', 'cod_suelo', 'sup_suelo', 'tipo_suelo']]


def agregar_por_predio(df, sufijo):
    """Agrupa líneas de suelo por predio usando operaciones vectorizadas."""

    # cod_comuna: tomar el primero por predio
    cod_comuna = df.groupby('clave_predio')['cod_comuna'].first()

    # Separar secano y riego
    secano = df[df['tipo_suelo'] == 'secano']
    riego = df[df['tipo_suelo'] == 'riego']

    # Superficies
    sup_secano = secano.groupby('clave_predio')['sup_suelo'].sum()
    sup_riego = riego.groupby('clave_predio')['sup_suelo'].sum()

    # Códigos como string concatenado
    codigos_secano = secano.groupby('clave_predio')['cod_suelo'].agg(
        lambda x: ','.join(sorted(x.unique()))
    )
    codigos_riego = riego.groupby('clave_predio')['cod_suelo'].agg(
        lambda x: ','.join(sorted(x.unique()))
    )

    # Armar resultado
    result = pd.DataFrame({'cod_comuna': cod_comuna})
    result[f'codigos_secano_{sufijo}'] = codigos_secano
    result[f'codigos_riego_{sufijo}'] = codigos_riego
    result[f'sup_secano_{sufijo}'] = sup_secano
    result[f'sup_riego_{sufijo}'] = sup_riego

    # Llenar NaN (predios que solo tienen un tipo)
    result[f'codigos_secano_{sufijo}'] = result[f'codigos_secano_{sufijo}'].fillna('')
    result[f'codigos_riego_{sufijo}'] = result[f'codigos_riego_{sufijo}'].fillna('')
    result[f'sup_secano_{sufijo}'] = result[f'sup_secano_{sufijo}'].fillna(0)
    result[f'sup_riego_{sufijo}'] = result[f'sup_riego_{sufijo}'].fillna(0)

    return result.reset_index()


def cargar_avaluos(filepath):
    """Carga un archivo de predios FWF y retorna avalúos para predios serie A."""
    print(f"  Cargando avalúos desde {filepath.name}...")

    chunks = pd.read_fwf(
        filepath,
        colspecs=PREDIOS_COLSPECS,
        names=PREDIOS_COLNAMES,
        encoding='latin-1',
        chunksize=300_000
    )

    dfs = []
    for chunk in chunks:
        chunk['cod_serie'] = chunk['cod_serie'].astype(str).str.strip().str.upper()
        chunk = chunk[chunk['cod_serie'] == 'A']

        chunk['comuna_actual'] = chunk['comuna_actual'].astype(str).str.strip().str.zfill(5)
        chunk['manzana_actual'] = pd.to_numeric(chunk['manzana_actual'], errors='coerce').astype('Int64')
        chunk['predio_actual'] = pd.to_numeric(chunk['predio_actual'], errors='coerce').astype('Int64')
        chunk['clave_predio'] = (
            chunk['comuna_actual'] + '-' +
            chunk['manzana_actual'].astype(str) + '-' +
            chunk['predio_actual'].astype(str)
        )
        chunk['avaluo_tot'] = pd.to_numeric(chunk['avaluo_tot'], errors='coerce')
        chunk['avaluo_ex'] = pd.to_numeric(chunk['avaluo_ex'], errors='coerce')
        chunk['cuota_trimestral'] = pd.to_numeric(chunk['cuota_trimestral'], errors='coerce')

        dfs.append(chunk[['clave_predio', 'avaluo_tot', 'avaluo_ex', 'cuota_trimestral']])

    df_avaluos = pd.concat(dfs, ignore_index=True)
    # Dedup: si hay duplicados de clave_predio, mantener el último
    df_avaluos = df_avaluos.drop_duplicates(subset=['clave_predio'], keep='last')

    print(f"    Predios serie A cargados: {len(df_avaluos):,}")
    return df_avaluos


def comparar_periodos(df_2019, df_2024):
    """Compara agregados por predio entre 2019 y 2024. Detecta cambios secano->riego."""

    # Inner join: predios que existen en ambos periodos
    df = df_2019.merge(df_2024, on='clave_predio', suffixes=('_2019', '_2024'))

    # Usar cod_comuna del 2024 (más reciente)
    if 'cod_comuna_2024' in df.columns:
        df['cod_comuna'] = df['cod_comuna_2024']
        df = df.drop(columns=['cod_comuna_2019', 'cod_comuna_2024'])
    elif 'cod_comuna' not in df.columns:
        df['cod_comuna'] = df_2024.set_index('clave_predio')['cod_comuna']

    print(f"\n  Predios en ambos periodos: {len(df):,}")

    # Calcular deltas
    df['delta_riego'] = df['sup_riego_2024'] - df['sup_riego_2019']
    df['delta_secano'] = df['sup_secano_2024'] - df['sup_secano_2019']

    # Requisito base: tenía secano en 2019, perdió secano Y ganó riego
    base = (
        (df['sup_secano_2019'] > 0) &
        (df['delta_secano'] < 0) &
        (df['delta_riego'] > 0)
    )

    # Clasificar tipo de cambio
    # Caso 1: No tenía riego en 2019 → superficie secana se convirtió a riego
    secano_a_riego = base & (df['sup_riego_2019'] == 0)
    # Caso 2: Ya tenía riego, pero ganó más a costa de secano
    secano_parcial_a_riego = base & (df['sup_riego_2019'] > 0)

    df['tipo_cambio'] = ''
    df.loc[secano_a_riego, 'tipo_cambio'] = 'secano_a_riego'
    df.loc[secano_parcial_a_riego, 'tipo_cambio'] = 'secano_parcial_a_riego'

    # Filtrar solo los que tuvieron conversión real de secano a riego
    df_cambios = df[df['tipo_cambio'] != ''].copy()

    print(f"  Predios con conversion secano->riego: {len(df_cambios):,}")
    print(f"    - secano_a_riego: {(df_cambios['tipo_cambio'] == 'secano_a_riego').sum():,}")
    print(f"    - secano_parcial_a_riego: {(df_cambios['tipo_cambio'] == 'secano_parcial_a_riego').sum():,}")

    return df_cambios


def generar_reportes(df_cambios, output_dir):
    """Genera Excel de detalle y CSV de resumen."""

    # --- Convertir avalúos a enteros ---
    for col in ['avaluo_tot_2019', 'avaluo_tot_2024', 'delta_avaluo',
                'cuota_trimestral_2019', 'cuota_trimestral_2024']:
        if col in df_cambios.columns:
            df_cambios[col] = pd.to_numeric(df_cambios[col], errors='coerce').astype('Int64')

    # --- Reporte detallado (Excel) ---
    cols_detalle = [
        'clave_predio', 'cod_comuna', 'nombre_comuna', 'rut', 'nombre_propietario',
        'codigos_secano_2019', 'codigos_riego_2019', 'sup_secano_2019', 'sup_riego_2019',
        'codigos_secano_2024', 'codigos_riego_2024', 'sup_secano_2024', 'sup_riego_2024',
        'delta_riego', 'delta_secano', 'tipo_cambio',
        'avaluo_tot_2019', 'avaluo_tot_2024', 'delta_avaluo', 'pct_cambio_avaluo',
        'cuota_trimestral_2019', 'cuota_trimestral_2024'
    ]
    cols_out = [c for c in cols_detalle if c in df_cambios.columns]
    path_detalle = output_dir / 'reporte_cambio_suelo_secano_riego.xlsx'
    df_out = df_cambios[cols_out]
    with pd.ExcelWriter(path_detalle, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='Detalle')
        ws = writer.sheets['Detalle']
        # Formato número para columnas de avalúo
        from openpyxl.utils import get_column_letter
        for col_name in ['avaluo_tot_2019', 'avaluo_tot_2024', 'delta_avaluo',
                         'cuota_trimestral_2019', 'cuota_trimestral_2024']:
            if col_name in cols_out:
                col_idx = cols_out.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df_out) + 2):
                    ws[f'{col_letter}{row}'].number_format = '#,##0'
        # Formato porcentaje
        if 'pct_cambio_avaluo' in cols_out:
            col_idx = cols_out.index('pct_cambio_avaluo') + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, len(df_out) + 2):
                ws[f'{col_letter}{row}'].number_format = '#,##0.0'
        # Auto-ajustar ancho de columnas
        for i, col_name in enumerate(cols_out, 1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[col_letter].width = max_len + 2
    print(f"\n  Reporte detallado: {path_detalle}")
    print(f"    Filas: {len(df_cambios):,}")

    # --- Resumen ---
    resumen_rows = []

    # Totales
    resumen_rows.append({
        'metrica': 'total_predios_con_cambio',
        'valor': len(df_cambios)
    })
    resumen_rows.append({
        'metrica': 'hectareas_riego_ganadas',
        'valor': df_cambios['delta_riego'].sum()
    })
    resumen_rows.append({
        'metrica': 'hectareas_secano_perdidas',
        'valor': abs(df_cambios['delta_secano'].sum())
    })

    # Por tipo de cambio
    for tipo in ['secano_a_riego', 'secano_parcial_a_riego']:
        sub = df_cambios[df_cambios['tipo_cambio'] == tipo]
        resumen_rows.append({
            'metrica': f'{tipo}_predios',
            'valor': len(sub)
        })
        resumen_rows.append({
            'metrica': f'{tipo}_hectareas_riego_ganadas',
            'valor': sub['delta_riego'].sum()
        })

    # Por comuna (top 20)
    por_comuna = df_cambios.groupby('cod_comuna').agg(
        predios=('clave_predio', 'count'),
        hect_riego_ganadas=('delta_riego', 'sum')
    ).sort_values('predios', ascending=False).head(20)

    for cod, row in por_comuna.iterrows():
        nombre = ''
        if 'nombre_comuna' in df_cambios.columns:
            match = df_cambios.loc[df_cambios['cod_comuna'] == cod, 'nombre_comuna']
            if len(match) > 0:
                nombre = match.iloc[0]
        resumen_rows.append({
            'metrica': f'comuna_{cod}_{nombre}',
            'valor': f"{int(row['predios'])} predios, {row['hect_riego_ganadas']:.0f} hect"
        })

    df_resumen = pd.DataFrame(resumen_rows)
    path_resumen = output_dir / 'resumen_cambio_suelo.csv'
    df_resumen.to_csv(path_resumen, index=False, encoding='utf-8-sig')
    print(f"  Resumen: {path_resumen}")


def main():
    print("=" * 60)
    print("DETECCION DE CAMBIO DE SUELO SECANO -> RIEGO (2019 vs 2024)")
    print("=" * 60)

    # Paso 1: Cargar líneas de suelo
    print("\n[1/5] Cargando construcciones agrícolas 2019...")
    df_suelo_2019 = cargar_suelos(ARCHIVO_AL_2019)

    print("\n[2/5] Cargando construcciones agrícolas 2024...")
    df_suelo_2024 = cargar_suelos(ARCHIVO_AL_2024)

    # Paso 2: Agregar por predio
    print("\n[3/5] Agregando por predio...")
    agg_2019 = agregar_por_predio(df_suelo_2019, '2019')
    agg_2024 = agregar_por_predio(df_suelo_2024, '2024')

    print(f"  Predios 2019: {len(agg_2019):,}")
    print(f"  Predios 2024: {len(agg_2024):,}")

    # Paso 3: Comparar periodos
    print("\n[4/5] Comparando periodos...")
    df_cambios = comparar_periodos(agg_2019, agg_2024)

    if len(df_cambios) == 0:
        print("\n  No se detectaron cambios de secano a riego.")
        return

    # Paso 4: Cruzar con avalúos (desde archivos locales)
    print("\n[5/5] Cruzando con avalúos...")
    avaluos_2019 = cargar_avaluos(ARCHIVO_PREDIOS_2019)
    avaluos_2024 = cargar_avaluos(ARCHIVO_PREDIOS_2024)

    df_cambios = df_cambios.merge(
        avaluos_2019.rename(columns={
            'avaluo_tot': 'avaluo_tot_2019', 'avaluo_ex': 'avaluo_ex_2019',
            'cuota_trimestral': 'cuota_trimestral_2019'
        }),
        on='clave_predio', how='left'
    )
    df_cambios = df_cambios.merge(
        avaluos_2024.rename(columns={
            'avaluo_tot': 'avaluo_tot_2024', 'avaluo_ex': 'avaluo_ex_2024',
            'cuota_trimestral': 'cuota_trimestral_2024'
        }),
        on='clave_predio', how='left'
    )

    df_cambios['delta_avaluo'] = df_cambios['avaluo_tot_2024'] - df_cambios['avaluo_tot_2019']
    df_cambios['pct_cambio_avaluo'] = np.where(
        df_cambios['avaluo_tot_2019'] > 0,
        (df_cambios['delta_avaluo'] / df_cambios['avaluo_tot_2019'] * 100).round(1),
        np.nan
    )

    # Nombres de comunas (intentar obtener de DB, sino dejar vacío)
    try:
        import psycopg2
        from dotenv import load_dotenv
        import os
        load_dotenv()
        conn = psycopg2.connect(
            dbname=os.getenv('PGDATABASE'),
            user=os.getenv('PGUSER'),
            password=os.getenv('PGPASSWORD'),
            host=os.getenv('PGHOST'),
            port=os.getenv('PGPORT'),
            sslmode="require"
        )
        df_comunas = pd.read_sql("SELECT cod_comuna, nombre_comuna FROM public.comunas", conn)
        df_cambios = df_cambios.merge(df_comunas, on='cod_comuna', how='left')
        print(f"  Nombres de comunas cargados desde DB.")

        # Propietarios: rut y nombre
        claves = df_cambios['clave_predio'].tolist()
        placeholders = ','.join([f"'{c}'" for c in claves])
        query_prop = f"""
            SELECT DISTINCT ON (pr.clave_predio)
                pr.clave_predio, p.rut, p.nombre AS nombre_propietario
            FROM public.propietarios p
            JOIN public.predios pr ON pr.predio_id = p.predio_id
            WHERE pr.clave_predio IN ({placeholders})
            ORDER BY pr.clave_predio, p.created_at DESC
        """
        df_prop = pd.read_sql(query_prop, conn)
        conn.close()
        df_cambios = df_cambios.merge(df_prop, on='clave_predio', how='left')
        print(f"  Propietarios cargados: {df_prop['rut'].notna().sum()} de {len(df_cambios)}")
    except Exception as e:
        print(f"  No se pudieron cargar nombres de comunas: {e}")
        df_cambios['nombre_comuna'] = ''

    # Paso 5: Generar reportes
    print("\n--- Generando reportes ---")
    generar_reportes(df_cambios, OUTPUT_DIR)

    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)


if __name__ == "__main__":
    main()
